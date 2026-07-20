"""
PDF to Excel Extractor
Extracts header fields + line items from Uptick/Anglicare Scope of Works PDFs.

Usage:
    python pdf_to_excel.py <input.pdf> [output.xlsx]
    python pdf_to_excel.py <input.pdf> [output.xlsx] --debug
    python pdf_to_excel.py <input.pdf> [output.xlsx] --showpages
"""

import sys
import re
from pathlib import Path
from collections import defaultdict

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


# ── Regex patterns — header fields ───────────────────────────────────────────
RE_WORK_REQUEST = re.compile(r'Anglicare Work Request[:\s]+(\S+)', re.IGNORECASE)
RE_CATEGORY     = re.compile(r'Renovation Category[:\s]+(\S+)', re.IGNORECASE)
RE_SCHEDULE     = re.compile(r'Renovation FFFE Schedule[:\s]+(\S+)', re.IGNORECASE)
RE_SCHEME       = re.compile(r'FFFE Scheme[:\s]+(.+?)(?:\n|$)', re.IGNORECASE)

# ── Regex patterns — codes ────────────────────────────────────────────────────
RE_SOR_LINE         = re.compile(r'SOR Activity Code:\s*([A-Z]{2,6}-\d{2,3})\b', re.IGNORECASE)
RE_FFFE_LINE        = re.compile(r'FFFE Code:\s*([A-Z0-9][A-Z0-9\-]{3,})\b',    re.IGNORECASE)
# FIX #2: unit is now optional so bare quantities like "1" (no unit) still parse.
RE_QUANTITY         = re.compile(r'([\d.]+)(?:\s+([A-Z][A-Z0-9\-]*))?', re.IGNORECASE)
RE_FFFE_ITEM_INLINE = re.compile(r'^FFFE Item:\s*(.+)$', re.IGNORECASE)
RE_FFFE_ITEM_BLANK  = re.compile(r'^FFFE Item[:\s]*$',   re.IGNORECASE)

RE_SKIP = re.compile(
    r'^(SOR Activity Code|FFFE Code|FFFE Item|Quantity|Required Parts|'
    r'Non-SoR Activities|Renovation Trade|Renovation –|Renovation Activity|'
    r'Page \d|Villa \d|Oran Park|Anglicare|Scope prepared|Overall Unit|'
    r'Renovation Category|Renovation FFFE|FFFE Scheme)',
    re.IGNORECASE
)

PAGE_HEIGHT = 1000


def is_bold(fontname: str) -> bool:
    return bool(re.search(r'bold|black|heavy|semibold', fontname, re.IGNORECASE))


def extract_header(full_text: str) -> dict:
    work_request = RE_WORK_REQUEST.search(full_text)
    category     = RE_CATEGORY.search(full_text)
    schedule     = RE_SCHEDULE.search(full_text)
    scheme       = RE_SCHEME.search(full_text)
    return {
        'Work Request':             work_request.group(1).strip() if work_request else '',
        'Renovation Category':      category.group(1).strip()     if category     else '',
        'Renovation FFFE Schedule': schedule.group(1).strip()     if schedule     else '',
        'FFFE Scheme':              scheme.group(1).strip()       if scheme       else '',
    }


def get_all_words(pdf_path: str):
    all_words = []
    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages):
            words = page.extract_words(extra_attrs=['fontname', 'size'])
            if not words:
                continue
            page_width = float(page.width)
            for w in words:
                w['page_num']   = page_num
                w['page_width'] = page_width
                w['doc_top']    = page_num * PAGE_HEIGHT + float(w['top'])
                all_words.append(w)
    return all_words


def parse_quantity(text: str):
    m = RE_QUANTITY.search(text)
    if m:
        return float(m.group(1)), (m.group(2).upper() if m.group(2) else '')
    return '', ''


def build_lines(word_list, bucket_size=4):
    groups = defaultdict(list)
    for w in word_list:
        bucket = round(w['doc_top'] / bucket_size) * bucket_size
        groups[bucket].append(w)
    lines = []
    for bucket in sorted(groups.keys()):
        lw       = sorted(groups[bucket], key=lambda w: float(w['x0']))
        text     = ' '.join(w['text'] for w in lw).strip()
        bold_cnt = sum(1 for w in lw if is_bold(w['fontname']))
        all_bold = bold_cnt == len(lw)
        if text:
            lines.append((bucket, text, all_bold))
    return lines


def build_right_blocks(pdf_path: str, page_cutoffs: dict, showpages: bool = False):
    """
    Extract SOR+FFFE pairs from the right column using PDF content stream order.

    Each SOR is paired with the next FFFE in the stream, skipping over any
    intervening non-code lines (e.g. page footers like "Page 18 of 22").
    If another SOR appears before a FFFE is found, the first SOR has no
    FFFE (Non-SoR item or genuinely missing code).
    """
    code_stream = []  # (kind, value): kind = 'sor' | 'fffe' | 'other'

    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages):
            cutoff      = page_cutoffs.get(page_num, float(page.width) * 0.55)
            page_width  = float(page.width)
            page_height = float(page.height)

            right_bbox = (cutoff, 0, page_width, page_height)
            right_page = page.crop(right_bbox)
            text       = right_page.extract_text(layout=False)

            if not text:
                continue

            if showpages:
                print(f"\n--- PAGE {page_num} RIGHT COLUMN (content stream) ---")
                for ln in text.split('\n'):
                    if ln.strip():
                        print(f"  {ln}")

            for line in text.split('\n'):
                line = line.strip()
                if not line:
                    continue
                sor_m = RE_SOR_LINE.search(line)
                if sor_m:
                    code_stream.append(('sor', sor_m.group(1)))
                    continue
                fffe_m = RE_FFFE_LINE.search(line)
                if fffe_m:
                    code_stream.append(('fffe', fffe_m.group(1)))
                    continue
                code_stream.append(('other', line))

    # Pair each SOR with the next FFFE, skipping over 'other' lines.
    # Stop looking (no FFFE) only if another SOR appears first.
    right_blocks = []
    seq          = 0
    i            = 0

    while i < len(code_stream):
        kind, value = code_stream[i]

        if kind == 'sor':
            sor_code  = value
            fffe_code = ''
            j = i + 1

            while j < len(code_stream) and code_stream[j][0] == 'other':
                j += 1

            if j < len(code_stream) and code_stream[j][0] == 'fffe':
                fffe_code = code_stream[j][1]
                i = j + 1
            else:
                i += 1

            right_blocks.append({
                'seq_idx':   seq,
                'sor_code':  sor_code,
                'fffe_code': fffe_code,
            })
            seq += 1
        else:
            i += 1

    return right_blocks


def extract_items(pdf_path: str, debug: bool = False, showpages: bool = False) -> list:
    full_text = ''
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                full_text += t + '\n'

    header    = extract_header(full_text)
    all_words = get_all_words(pdf_path)

    pages_words = defaultdict(list)
    for w in all_words:
        pages_words[w['page_num']].append(w)

    page_cutoffs = {}
    for page_num, pwords in pages_words.items():
        page_width  = pwords[0]['page_width']
        right_col_x = page_width
        for w in pwords:
            if re.match(r'^(SOR|FFFE)', w['text'], re.IGNORECASE):
                x0 = float(w['x0'])
                if x0 > page_width * 0.45:
                    right_col_x = min(right_col_x, x0)
        page_cutoffs[page_num] = (
            right_col_x * 0.95 if right_col_x < page_width
            else page_width * 0.55
        )

    left_words = [w for w in all_words
                  if float(w['x0']) < page_cutoffs[w['page_num']]]
    left_lines = build_lines(left_words)

    right_blocks = build_right_blocks(pdf_path, page_cutoffs, showpages=showpages)

    # ── Tag left-column lines ─────────────────────────────────────────────────
    tagged = []
    for doc_top, text, all_bold in left_lines:
        if re.match(r'^Quantity[:\s]*$', text, re.IGNORECASE):
            tagged.append((doc_top, text, all_bold, 'quantity_label'))
        elif re.match(r'^Quantity[:\s]', text, re.IGNORECASE):
            tagged.append((doc_top, text, all_bold, 'quantity'))
        elif RE_FFFE_ITEM_INLINE.match(text):
            m = RE_FFFE_ITEM_INLINE.match(text)
            tagged.append((doc_top, m.group(1).strip(), all_bold, 'fffe_item_value'))
        elif RE_FFFE_ITEM_BLANK.match(text):
            tagged.append((doc_top, text, all_bold, 'fffe_item_label'))
        elif RE_SKIP.match(text):
            tagged.append((doc_top, text, all_bold, 'meta'))
        elif all_bold:
            tagged.append((doc_top, text, all_bold, 'description'))
        else:
            tagged.append((doc_top, text, all_bold, 'value'))

    if debug:
        print('\n=== TAGGED LEFT COLUMN LINES ===')
        for doc_top, text, bold, role in tagged:
            print(f'  {doc_top:8.1f} [{role:20s}] {"[B]" if bold else "   "} {text}')
        print(f'\n=== RIGHT BLOCKS ({len(right_blocks)}) ===')
        for b in right_blocks:
            print(f'  [{b["seq_idx"]:3d}] SOR={b["sor_code"]:12s}  FFFE={b["fffe_code"]}')

    # ── Walk tagged lines to collect raw items ────────────────────────────────
    # Key fix: is_non_sor is determined by whether "Required Parts:" appears
    # in this item's own block (a direct structural marker on Non-SoR rows),
    # rather than tracking Non-SoR Activities/Renovation Trade section state,
    # which breaks when multiple section headings appear between Non-SoR items.
    raw_items = []
    i = 0

    while i < len(tagged):
        doc_top, text, bold, role = tagged[i]

        if role == 'description':
            item = {
                'desc':       text,
                'desc_top':   doc_top,
                'qty':        '',
                'unit':       '',
                'fffe_item':  '',
                'is_non_sor': False,
            }
            j = i + 1

            while j < len(tagged):
                t2, txt2, b2, r2 = tagged[j]

                if r2 == 'description':
                    break
                elif r2 == 'meta' and re.match(r'^Required Parts', txt2, re.IGNORECASE):
                    item['is_non_sor'] = True
                    j += 1
                elif r2 == 'quantity' and not item['qty']:
                    q, u = parse_quantity(txt2)
                    item['qty']  = q
                    item['unit'] = u
                    j += 1
                elif r2 == 'quantity_label' and not item['qty']:
                    if j - 1 >= i and tagged[j - 1][3] == 'value':
                        q, u = parse_quantity(tagged[j - 1][1])
                        item['qty']  = q
                        item['unit'] = u
                    j += 1
                elif r2 == 'fffe_item_value' and not item['fffe_item']:
                    item['fffe_item'] = txt2
                    j += 1
                    if item['qty']:
                        break
                elif r2 == 'fffe_item_label' and not item['fffe_item']:
                    if j + 1 < len(tagged):
                        t3, txt3, b3, r3 = tagged[j + 1]
                        if r3 == 'value' and not b3:
                            item['fffe_item'] = txt3
                            j += 2
                        else:
                            j += 1
                    else:
                        j += 1
                    if item['qty']:
                        break
                elif r2 == 'meta':
                    j += 1
                else:
                    j += 1

            raw_items.append(item)
            i = j
        else:
            i += 1

    # ── FIX #1 + Location tracking ──────────────────────────────────────────────
    # A real item always carries at least one signal: a quantity, an FFFE Item
    # value, or a Required Parts marker (Non-SoR). Bold section/room headings
    # (e.g. "Room 57 - Warrah Wing - Ensuite") never carry any of these,
    # regardless of property-naming convention, so this generalizes cleanly
    # without hardcoding any property-specific strings.
    #
    # Rather than just discarding those heading lines, track them as the
    # current "location" and carry it forward onto every real item that
    # follows, until the next heading changes it. Defaults to "Overall Unit
    # Activities" for items appearing before any room-specific heading.
    current_location = 'Overall Unit Activities'
    located_items     = []
    for r in raw_items:
        is_real = r['is_non_sor'] or r['qty'] != '' or r['fffe_item'] != ''
        if not is_real:
            current_location = r['desc'].strip()
            continue
        r['location'] = current_location
        located_items.append(r)
    raw_items = located_items

    # ── Pair descriptions with right blocks, skipping Non-SoR items ───────────
    items     = []
    block_idx = 0

    for raw in raw_items:
        if raw['is_non_sor']:
            sor_code, fffe_code = '', ''
        else:
            block = right_blocks[block_idx] if block_idx < len(right_blocks) else {'sor_code': '', 'fffe_code': ''}
            sor_code, fffe_code = block['sor_code'], block['fffe_code']
            block_idx += 1

        items.append({
            'Location':                 raw['location'],
            'Work Request':             header['Work Request'],
            'Renovation Category':      header['Renovation Category'],
            'Renovation FFFE Schedule': header['Renovation FFFE Schedule'],
            'FFFE Scheme':              header['FFFE Scheme'],
            'Description':              raw['desc'],
            'Quantity':                 raw['qty'],
            'Unit':                     raw['unit'],
            'SOR Activity Code':        sor_code,
            'FFFE Code':                fffe_code,
            'FFFE Item':                raw['fffe_item'],
            'Source PDF':               Path(pdf_path).name,
            # Matches costs.xlsx column E ("SOR Lookup"), e.g. 'RC-DEM-01'.
            'SOR Lookup Key':           f"{header['Renovation FFFE Schedule']}-{sor_code}" if sor_code else '',
        })

    return items


def style_header_row(ws):
    header_fill = PatternFill('solid', start_color='1F4E79')
    header_font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    border      = Border(bottom=Side(style='medium', color='FFFFFF'))
    for cell in ws[1]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = border


def style_rows(ws, start_row: int):
    light       = PatternFill('solid', start_color='DCE6F1')
    normal_font = Font(name='Arial', size=10)
    for row_idx, row in enumerate(ws.iter_rows(min_row=start_row, max_row=ws.max_row), start=0):
        fill = light if row_idx % 2 == 0 else None
        for cell in row:
            cell.font      = normal_font
            if fill:
                cell.fill  = fill
            cell.alignment = Alignment(vertical='center', wrap_text=True)


def set_column_widths(ws, col_widths: dict):
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width


def build_table(ws, num_cols: int, table_name: str):
    """Add a real Excel Table (ListObject) covering the sheet's data, so
    tools like Power Automate's Excel Online connector can run 'List rows' /
    'Update a row' against it."""
    last_row = ws.max_row
    last_col_letter = get_column_letter(num_cols)
    table_ref = f"A1:{last_col_letter}{last_row}"

    table = Table(displayName=table_name, ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def write_sheet(wb, sheet_title: str, headers: list, col_widths: dict,
                 items: list, table_name: str, is_active: bool = False):
    ws = wb.create_sheet(title=sheet_title) if not is_active else wb.active
    ws.title = sheet_title
    ws.row_dimensions[1].height = 30

    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)

    style_header_row(ws)
    ws.freeze_panes = 'A2'

    for item in items:
        ws.append([item.get(h, '') for h in headers])

    style_rows(ws, start_row=2)
    set_column_widths(ws, col_widths)
    build_table(ws, len(headers), table_name)
    return ws


def write_validation_sheet(wb, sheet_title: str, headers: list, col_widths: dict,
                            items: list, table_name: str):
    """
    Same shape as write_sheet, except 'CBC Qty' is always left blank (to be
    filled in by hand later) and 'CBC Qty Total ($)' is written as a live
    Excel formula (FINAL * CBC Qty for that row), so it recalculates
    automatically once someone fills in CBC Qty - rather than a fixed value
    computed once at generation time.
    """
    ws = wb.create_sheet(title=sheet_title)
    ws.row_dimensions[1].height = 30

    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)
    style_header_row(ws)
    ws.freeze_panes = 'A2'

    cbc_qty_col   = get_column_letter(headers.index('CBC Qty') + 1)
    final_col     = get_column_letter(headers.index('FINAL') + 1)
    cbc_total_idx = headers.index('CBC Qty Total ($)') + 1

    for row_offset, item in enumerate(items):
        row = row_offset + 2
        for col_idx, h in enumerate(headers, start=1):
            if h == 'CBC Qty':
                continue  # left blank - filled in by hand later
            elif h == 'CBC Qty Total ($)':
                ws.cell(row=row, column=col_idx,
                        value=f"={final_col}{row}*{cbc_qty_col}{row}")
            else:
                ws.cell(row=row, column=col_idx, value=item.get(h, ''))

    style_rows(ws, start_row=2)
    set_column_widths(ws, col_widths)
    build_table(ws, len(headers), table_name)
    return ws


def write_grouped_sheet(wb, sheet_title: str, headers: list, col_widths: dict, items: list):
    """
    Same columns as a normal sheet, but rows are grouped by the first 3
    characters of 'SOR Activity Code' (e.g. 'DEM', 'PAT', 'FLT'). Items with
    no/short SOR code (Non-SoR activities) are grouped under 'OTHER'.

    Each group gets a bold, merged section-header row, followed by its item
    rows. Item rows carry an Excel outline level so Excel's native +/-
    row-grouping controls can collapse/expand each group; the header row
    itself is the summary row (outlinePr.summaryBelow = False puts the
    collapse control at the header, not below the detail rows).

    Not turned into an Excel Table, since Tables require one contiguous
    header row immediately followed by data — the repeated group-header rows
    here don't fit that shape. This is a human-readable grouped view, not a
    machine-queried one.
    """
    ws = wb.create_sheet(title=sheet_title)
    ws.row_dimensions[1].height = 30

    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)
    style_header_row(ws)
    ws.freeze_panes = 'A2'

    # Group items by first 3 chars of SOR Activity Code (upper), 'OTHER' for
    # blank/short codes (e.g. Non-SoR activities).
    groups = defaultdict(list)
    for item in items:
        code = (item.get('SOR Activity Code') or '').strip().upper()
        prefix = code[:3] if len(code) >= 3 else 'OTHER'
        groups[prefix].append(item)

    last_col_letter = get_column_letter(len(headers))
    group_fill = PatternFill('solid', start_color='B8CCE4')
    group_font = Font(bold=True, name='Arial', size=10)

    row_idx = 2
    for prefix in sorted(groups.keys()):
        group_items = groups[prefix]

        ws.cell(row=row_idx, column=1, value=f"{prefix} ({len(group_items)} item{'s' if len(group_items) != 1 else ''})")
        ws.merge_cells(f"A{row_idx}:{last_col_letter}{row_idx}")
        header_cell = ws.cell(row=row_idx, column=1)
        header_cell.fill = group_fill
        header_cell.font = group_font
        header_cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row_idx].outlineLevel = 0
        row_idx += 1

        for item in group_items:
            for col_idx, h in enumerate(headers, start=1):
                ws.cell(row=row_idx, column=col_idx, value=item.get(h, ''))
            ws.row_dimensions[row_idx].outlineLevel = 1
            row_idx += 1

    style_rows(ws, start_row=2)
    set_column_widths(ws, col_widths)

    # Collapse control sits on the group-header row (summary above detail),
    # matching this layout (header row, then its items below).
    ws.sheet_properties.outlinePr.summaryBelow = False
    return ws


def write_to_excel(items: list, output_path: str):
    headers = [
        'Location',
        'Work Request', 'Renovation Category', 'Renovation FFFE Schedule',
        'FFFE Scheme', 'Description', 'Quantity', 'Unit',
        'SOR Activity Code', 'FFFE Code', 'FFFE Item', 'Source PDF',
        'SOR Lookup Key',
        'Sub-contractor COST ($)', 'Sub-contractor Total ($)',
        'Anglicare Cost', 'Anglicare Total ($)',
        'CBC Cost', 'CBC Total ($)',
        'FINAL', 'FINAL Total ($)',
    ]
    col_widths = {
        'A': 26,
        'B': 18, 'C': 22, 'D': 26, 'E': 18,
        'F': 60, 'G': 12, 'H': 12, 'I': 20,
        'J': 25, 'K': 30, 'L': 30, 'M': 20,
        'N': 20, 'O': 20, 'P': 18, 'Q': 18,
        'R': 14, 'S': 16, 'T': 14, 'U': 16,
    }

    # Second/third tab: columns A-O only (Location through Sub-contractor
    # Total), excluding the Anglicare/CBC/FINAL cost columns.
    subcontractor_headers    = headers[:15]
    subcontractor_col_widths = {k: v for k, v in col_widths.items() if k in
                                 ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H',
                                  'I', 'J', 'K', 'L', 'M', 'N', 'O']}

    # Third tab: same as the second, plus one extra column concatenating
    # columns A-M (Location through SOR Lookup Key) separated by ' / '.
    combined_source_headers = headers[:13]  # A-M: Location .. SOR Lookup Key
    grouped_headers         = subcontractor_headers + ['Combined Reference']
    grouped_col_widths      = dict(subcontractor_col_widths)
    grouped_col_widths['P'] = 70

    grouped_items = []
    for item in items:
        combined = ' / '.join(str(item.get(h, '')) for h in combined_source_headers)
        new_item = dict(item)
        new_item['Combined Reference'] = combined
        grouped_items.append(new_item)

    # Name -> width lookup (built once from the original letter-keyed dict),
    # so new sheets with rearranged columns can look up widths by header name
    # instead of recomputing letters by hand.
    name_to_width = {h: col_widths[get_column_letter(i + 1)] for i, h in enumerate(headers)}
    name_to_width['CBC Qty']            = 12
    name_to_width['CBC Qty Total ($)']  = 18
    name_to_width['Combined Reference'] = 70

    # Fourth tab "Validation": same as tab 1, minus Sub-contractor COST/Total,
    # with 'CBC Qty' inserted right after Quantity (blank - filled in by hand
    # later) and 'CBC Qty Total ($)' added as the last column, as a live
    # formula (= FINAL * CBC Qty) so it recalculates once CBC Qty is filled in.
    validation_base = [h for h in headers
                        if h not in ('Sub-contractor COST ($)', 'Sub-contractor Total ($)')]
    qty_idx = validation_base.index('Quantity')
    validation_headers = (
        validation_base[:qty_idx + 1] + ['CBC Qty'] + validation_base[qty_idx + 1:]
        + ['CBC Qty Total ($)']
    )
    validation_col_widths = {
        get_column_letter(i + 1): name_to_width.get(h, 15)
        for i, h in enumerate(validation_headers)
    }

    wb = Workbook()
    write_sheet(wb, 'Extracted Data', headers, col_widths, items,
                table_name='LineItems', is_active=True)
    write_sheet(wb, 'Subcontractor Copy', subcontractor_headers,
                subcontractor_col_widths, items,
                table_name='SubcontractorView')
    write_grouped_sheet(wb, 'Grouped by SOR', grouped_headers,
                        grouped_col_widths, grouped_items)
    write_validation_sheet(wb, 'Validation', validation_headers,
                           validation_col_widths, items,
                           table_name='ValidationData')

    wb.save(output_path)


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    pdf_path = sys.argv[1]
    if not Path(pdf_path).exists():
        print(f"Error: File not found — {pdf_path}")
        sys.exit(1)

    debug     = '--debug'     in sys.argv
    showpages = '--showpages' in sys.argv
    output_path = next(
        (a for a in sys.argv[2:] if not a.startswith('--')),
        str(Path(pdf_path).with_suffix('.xlsx'))
    )

    print(f"Reading:  {pdf_path}")
    items = extract_items(pdf_path, debug=debug, showpages=showpages)

    if not items:
        print("No items extracted.")
        sys.exit(1)

    h = items[0]
    print(f"  Work Request:             {h['Work Request']}")
    print(f"  Renovation Category:      {h['Renovation Category']}")
    print(f"  Renovation FFFE Schedule: {h['Renovation FFFE Schedule']}")
    print(f"  FFFE Scheme:              {h['FFFE Scheme']}")
    print(f"  Line items extracted:     {len(items)}")

    write_to_excel(items, output_path)
    print(f"Saved to: {output_path}")


# NOTE: main()/CLI entry point kept above for standalone/local testing
# (`python pdf_extractor.py <input.pdf>`). It is NOT used by the Azure
# Function wrapper in function_app.py, which calls extract_items() and
# write_to_excel() directly.
if __name__ == '__main__':
    main()
